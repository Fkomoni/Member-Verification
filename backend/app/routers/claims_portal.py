"""
Claims Portal endpoints — for the claims team to track, review, and export claims.

All endpoints require agent auth with claims_officer or admin role.
"""

import io
import logging
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.core.database import get_db
from app.core.deps import get_current_agent, require_role
from app.models.models import (
    Agent,
    AuthorizationCode,
    ClaimAuditLog,
    ClaimServiceLine,
    ReimbursementClaim,
)
from app.schemas.schemas import (
    AuditLogResponse,
    ClaimDetailResponse,
    ClaimsListResponse,
    ReimbursementClaimResponse,
    ServiceLineItem,
    UpdateClaimStatusRequest,
)
from app.services import audit_service

log = logging.getLogger(__name__)

router = APIRouter(prefix="/claims-portal", tags=["claims-portal"])


# ── Claims List ──────────────────────────────────────────────


@router.get("/claims", response_model=ClaimsListResponse)
def list_claims(
    status_filter: str | None = Query(None, alias="status"),
    search: str | None = Query(None),
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """List all reimbursement claims with optional filters."""
    query = db.query(ReimbursementClaim).options(
        joinedload(ReimbursementClaim.authorization_code)
    )

    if status_filter:
        query = query.filter(ReimbursementClaim.status == status_filter)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (ReimbursementClaim.claim_ref.ilike(search_term))
            | (ReimbursementClaim.enrollee_id.ilike(search_term))
            | (ReimbursementClaim.member_name.ilike(search_term))
        )

    total = query.count()
    claims = query.order_by(ReimbursementClaim.created_at.desc()).offset(skip).limit(limit).all()

    return ClaimsListResponse(
        claims=[
            ReimbursementClaimResponse(
                claim_id=c.claim_id,
                claim_ref=c.claim_ref,
                enrollee_id=c.enrollee_id,
                member_name=c.member_name,
                hospital_name=c.hospital_name,
                claim_amount=c.claim_amount,
                status=c.status,
                authorization_code=c.authorization_code.code if c.authorization_code else None,
                created_at=c.created_at,
            )
            for c in claims
        ],
        total=total,
    )


# ── Claim Detail ─────────────────────────────────────────────


@router.get("/claims/{claim_id}", response_model=ClaimDetailResponse)
def get_claim_detail(
    claim_id: uuid.UUID,
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """Get full detail of a specific claim."""
    claim = (
        db.query(ReimbursementClaim)
        .options(
            joinedload(ReimbursementClaim.authorization_code),
            joinedload(ReimbursementClaim.service_lines),
        )
        .filter(ReimbursementClaim.claim_id == claim_id)
        .first()
    )

    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")

    auth_code = claim.authorization_code

    return ClaimDetailResponse(
        claim_id=claim.claim_id,
        claim_ref=claim.claim_ref,
        enrollee_id=claim.enrollee_id,
        member_name=claim.member_name,
        member_phone=claim.member_phone,
        hospital_name=claim.hospital_name,
        visit_date=claim.visit_date,
        reason_for_visit=claim.reason_for_visit,
        reimbursement_reason=claim.reimbursement_reason,
        claim_amount=claim.claim_amount,
        medications=claim.medications,
        lab_investigations=claim.lab_investigations,
        comments=claim.comments,
        bank_name=claim.bank_name,
        account_number=claim.account_number,
        account_name=claim.account_name,
        status=claim.status,
        approved_amount=claim.approved_amount,
        reviewer_notes=claim.reviewer_notes,
        authorization_code=auth_code.code if auth_code else None,
        agent_name=auth_code.agent_name if auth_code else None,
        service_lines=[
            ServiceLineItem(
                service_name=sl.service_name,
                quantity=sl.quantity,
                unit_price=sl.unit_price,
            )
            for sl in claim.service_lines
        ],
        created_at=claim.created_at,
        updated_at=claim.updated_at,
    )


# ── Update Claim Status ──────────────────────────────────────


@router.patch("/claims/{claim_id}/status")
def update_claim_status(
    claim_id: uuid.UUID,
    body: UpdateClaimStatusRequest,
    request: Request,
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """Update the status of a claim (review workflow)."""
    claim = db.query(ReimbursementClaim).filter(
        ReimbursementClaim.claim_id == claim_id
    ).first()

    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")

    old_status = claim.status
    claim.status = body.status
    claim.reviewer_id = agent.agent_id

    if body.approved_amount is not None:
        claim.approved_amount = float(body.approved_amount)
    if body.reviewer_notes is not None:
        claim.reviewer_notes = body.reviewer_notes

    db.commit()

    # Audit
    audit_service.log_action(
        db,
        entity_type="claim",
        entity_id=str(claim.claim_id),
        action="status_changed",
        actor_type="agent",
        actor_id=str(agent.agent_id),
        details={
            "claim_ref": claim.claim_ref,
            "old_status": old_status,
            "new_status": body.status,
            "approved_amount": float(body.approved_amount) if body.approved_amount else None,
            "reviewer_notes": body.reviewer_notes,
        },
        ip_address=request.client.host if request.client else None,
    )

    return {
        "success": True,
        "claim_ref": claim.claim_ref,
        "old_status": old_status,
        "new_status": body.status,
    }


# ── Claim Timeline (Audit Trail) ─────────────────────────────


@router.get("/claims/{claim_id}/timeline", response_model=list[AuditLogResponse])
def get_claim_timeline(
    claim_id: uuid.UUID,
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """Get the full audit trail / timeline for a claim."""
    # Get audit logs for the claim itself
    claim_logs = audit_service.get_audit_trail(
        db, entity_type="claim", entity_id=str(claim_id)
    )

    # Also get related authorization code logs
    claim = db.query(ReimbursementClaim).filter(
        ReimbursementClaim.claim_id == claim_id
    ).first()

    code_logs = []
    if claim and claim.authorization_code_id:
        code_logs = audit_service.get_audit_trail(
            db, entity_type="authorization_code", entity_id=str(claim.authorization_code_id)
        )

    all_logs = sorted(claim_logs + code_logs, key=lambda x: x.created_at, reverse=True)

    return [
        AuditLogResponse(
            id=l.id,
            entity_type=l.entity_type,
            entity_id=l.entity_id,
            action=l.action,
            actor_type=l.actor_type,
            actor_id=l.actor_id,
            details=l.details,
            ip_address=l.ip_address,
            created_at=l.created_at,
        )
        for l in all_logs
    ]


# ── Stats / Summary ──────────────────────────────────────────


@router.get("/stats")
def get_claims_stats(
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """Get summary statistics for the claims dashboard."""
    total = db.query(func.count(ReimbursementClaim.claim_id)).scalar() or 0
    by_status = (
        db.query(ReimbursementClaim.status, func.count(ReimbursementClaim.claim_id))
        .group_by(ReimbursementClaim.status)
        .all()
    )
    total_amount = (
        db.query(func.sum(ReimbursementClaim.claim_amount)).scalar() or 0
    )
    total_approved = (
        db.query(func.sum(ReimbursementClaim.approved_amount))
        .filter(ReimbursementClaim.status.in_(["approved", "payment_processing", "paid"]))
        .scalar() or 0
    )

    return {
        "total_claims": total,
        "total_claim_amount": float(total_amount),
        "total_approved_amount": float(total_approved),
        "by_status": {s: c for s, c in by_status},
    }


# ── Excel Export ─────────────────────────────────────────────


@router.get("/export")
def export_claims_excel(
    status_filter: str | None = Query(None, alias="status"),
    db: Session = Depends(get_db),
    agent: Agent = Depends(require_role("claims_officer", "admin")),
):
    """
    Export claims to Excel (.xlsx).
    Includes: auth code, member ID, member name, claim amount,
    approved amount, agent name, agent ID, notes, status, dates.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Add it to requirements.txt.",
        )

    query = (
        db.query(ReimbursementClaim)
        .options(joinedload(ReimbursementClaim.authorization_code))
        .order_by(ReimbursementClaim.created_at.desc())
    )

    if status_filter:
        query = query.filter(ReimbursementClaim.status == status_filter)

    claims = query.all()

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reimbursement Claims"

    # Header styles
    header_fill = PatternFill(start_color="C61531", end_color="C61531", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    headers = [
        "Claim Ref",
        "Authorization Code",
        "Member ID",
        "Member Name",
        "Phone",
        "Hospital",
        "Visit Date",
        "Claim Amount",
        "Approved Amount",
        "Agent Name",
        "Agent ID",
        "Status",
        "Reimbursement Reason",
        "Reason for Visit",
        "Medications",
        "Lab Investigations",
        "Bank",
        "Account No",
        "Account Name",
        "Notes",
        "Reviewer Notes",
        "Submitted",
        "Updated",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, c in enumerate(claims, 2):
        auth = c.authorization_code
        ws.cell(row=row, column=1, value=c.claim_ref)
        ws.cell(row=row, column=2, value=auth.code if auth else "")
        ws.cell(row=row, column=3, value=c.enrollee_id)
        ws.cell(row=row, column=4, value=c.member_name)
        ws.cell(row=row, column=5, value=c.member_phone)
        ws.cell(row=row, column=6, value=c.hospital_name)
        ws.cell(row=row, column=7, value=str(c.visit_date) if c.visit_date else "")
        ws.cell(row=row, column=8, value=float(c.claim_amount))
        ws.cell(row=row, column=9, value=float(c.approved_amount) if c.approved_amount else "")
        ws.cell(row=row, column=10, value=auth.agent_name if auth else "")
        ws.cell(row=row, column=11, value=str(auth.agent_id) if auth else "")
        ws.cell(row=row, column=12, value=c.status)
        ws.cell(row=row, column=13, value=c.reimbursement_reason)
        ws.cell(row=row, column=14, value=c.reason_for_visit)
        ws.cell(row=row, column=15, value=c.medications or "")
        ws.cell(row=row, column=16, value=c.lab_investigations or "")
        ws.cell(row=row, column=17, value=c.bank_name)
        ws.cell(row=row, column=18, value=c.account_number)
        ws.cell(row=row, column=19, value=c.account_name)
        ws.cell(row=row, column=20, value=auth.notes if auth else "")
        ws.cell(row=row, column=21, value=c.reviewer_notes or "")
        ws.cell(row=row, column=22, value=c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "")
        ws.cell(row=row, column=23, value=c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "")

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"claims_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
